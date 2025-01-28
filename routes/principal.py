from flask import Blueprint, render_template, request, redirect, session, flash, jsonify, url_for, send_file
import pandas as pd
import os
from io import BytesIO
from flask_mysqldb import MySQL
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime

mysql = MySQL()

principal_bp = Blueprint('index', __name__, template_folder='templates')

# Ruta para el index 
@principal_bp.route("/principal")
def index():
    if session.get("logueado"):
        return render_template("principal.html")
    else:
        return redirect("/")
    
# Ruta de cargar turnos y transformar
@principal_bp.route('/cargar-turnos', methods=['POST'])
def cargar_turnos():
    if 'file' not in request.files:
        flash('No se envió ningún archivo', 'danger')
        return redirect("/principal")

    file = request.files['file']
    if file.filename == '':
        flash('No seleccionaste ningún archivo', 'danger')
        return redirect("/principal")

    if file:
        try:
            # Leer el archivo cargado
            df = pd.read_excel(file, dtype={'DNI_MEDICO': str, 'CMP': str})

            # Asegurar que las columnas necesarias existen
            required_columns = ['DNI_MEDICO', 'CMP', 'PROFESIONAL', 'ESPECIALIDAD', 'AREA', 'GRUPO_OCUPACIONAL', 'FECHA_PROGRAMACION', 'HOR_INICIO', 'HOR_FIN']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                flash(f'El archivo no cumple con el formato requerido. Faltan las siguientes columnas: {', '.join(missing_columns)}', 'danger')
                return redirect("/principal")

            # Convertir la columna de fecha a formato datetime y formatearla sin horas ni minutos
            df['FECHA_PROGRAMACION'] = pd.to_datetime(df['FECHA_PROGRAMACION'], errors='coerce').dt.strftime('%d/%m/%Y')
            
            # Asegurar que CMP tenga exactamente 6 dígitos, rellenando con ceros a la izquierda
            df['CMP'] = df['CMP'].str.zfill(6)
            
            # Asegurar que DNI_MEDICO tenga exactamente 8 dígitos, rellenando con ceros a la izquierda
            df['DNI_MEDICO'] = df['DNI_MEDICO'].str.zfill(8)
            
            # Agregar las columnas desde J con formato DIAxx_INGRESO y DIAxx_SALIDA
            for dia in range(1, 32):
                df[f'DIA{dia:02d}_INGRESO'] = ''  # Agregar columna vacía
                df[f'DIA{dia:02d}_SALIDA'] = ''   # Agregar columna vacía

            # Llenar las columnas con los valores de ingreso y salida
            for index, row in df.iterrows():
                if pd.notna(row['FECHA_PROGRAMACION']):
                    dia = int(row['FECHA_PROGRAMACION'][:2])  # Extraer el día del mes
                    df.at[index, f'DIA{dia:02d}_INGRESO'] = row['HOR_INICIO']
                    df.at[index, f'DIA{dia:02d}_SALIDA'] = row['HOR_FIN']

            # Reordenar las columnas en el orden correcto
            column_order = required_columns + [col for dia in range(1, 32) for col in (f'DIA{dia:02d}_INGRESO', f'DIA{dia:02d}_SALIDA')]
            df = df[column_order]

            # Crear un archivo Excel en memoria con los cambios
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Hoja1')
            output.seek(0)

            # Aplicar estilos al archivo generado
            wb = load_workbook(output)
            ws = wb.active

            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Fondo amarillo
            font = Font(color='FF0000')  # Texto rojo
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            
            # Aplicar estilo a las columnas especificadas
            columns_to_style = ['GRUPO_OCUPACIONAL', 'FECHA_PROGRAMACION', 'HOR_INICIO', 'HOR_FIN']
            for col in columns_to_style:
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1  # Obtener índice de columna (1-based)
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.fill = fill
                            cell.font = font
                            cell.border = border
            
            # Aplicar bordes a toda la información
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border

            # Obtener la fecha actual para el nombre del archivo
            fecha_actual = datetime.now().strftime('%Y%m%d')
            nombre_archivo = f'Planilla_Modificada_{fecha_actual}.xlsx'

            # Guardar el archivo con los estilos aplicados
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Retornar el archivo modificado como respuesta
            return send_file(output, 
                             as_attachment=True, 
                             download_name=nombre_archivo, 
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            flash(f'Error al procesar el archivo: {e}', 'danger')
            return redirect("/principal")